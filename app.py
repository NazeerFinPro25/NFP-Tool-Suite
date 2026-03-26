import streamlit as st
import pandas as pd
import datetime
import random
import io
import urllib.parse
import base64
import os
import re
import pdfplumber
from PIL import Image as PILImage
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CONFIGURATION & CSS
# ==========================================
st.set_page_config(
    page_title="NFP | Attendance Generator",
    page_icon="📊",
    layout="wide"
)

# --- BRAND CONSTANTS ---
BRAND_NAME = "NazeerFinPro"
BRAND_SHORT = "NFP"
BRAND_COLOR = "#003366" # Professional Navy Blue

# --- CSS FOR PROFESSIONAL LOOK ---
st.markdown(f"""
    <style>
    .main {{
        padding-top: 0rem;
    }}
    .stButton>button {{
        width: 100%;
        background-color: {BRAND_COLOR};
        color: white;
        font-weight: bold;
        border-radius: 5px;
        height: 3em; /* Consistent height */
    }}
    .stButton>button:hover {{
        background-color: #002244;
        color: white;
    }}
    .brand-header {{
        color: {BRAND_COLOR};
        font-size: 36px;
        font-weight: bold;
        text-align: left;
        margin-bottom: 0px;
        line-height: 1.2;
        padding-top: 10px;
    }}
    .brand-sub {{
        color: #666;
        font-size: 16px;
        text-align: left;
        margin-top: 0px;
        margin-bottom: 20px;
    }}
    .footer {{
        position: fixed;
        left: 0;
        bottom: 0;
        width: 100%;
        background-color: #f1f1f1;
        color: #333;
        text-align: center;
        padding: 10px;
        font-size: 12px;
        z-index: 999;
    }}
    .status-card {{ 
        padding: 20px; 
        border-radius: 10px; 
        background-color: #ffffff; 
        border: 1px solid #e0e0e0; 
        box-shadow: 2px 2px 5px rgba(0,0,0,0.05); 
        margin-bottom: 20px;
    }}
    div[data-testid="column"] {{
        display: flex;
        align-items: center; 
    }}
    /* Metrics Styling */
    div[data-testid="stMetricValue"] {{
        font-size: 24px;
        color: {BRAND_COLOR};
    }}
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# 2. HELPER FUNCTIONS (ALL)
# ==========================================

# --- A. ATTENDANCE HELPERS ---
def create_natural_time(year, month, base_hour, is_arrival):
    """Generates a natural-looking time string."""
    if is_arrival:
        minute = random.randint(-5, 10)
    else:
        minute = random.randint(0, 10)
    
    try:
        base_time = datetime.datetime(year, month, 1, base_hour, 0)
        final_time = base_time + datetime.timedelta(minutes=minute)
        return final_time.strftime("%H:%M")
    except ValueError:
        return "00:00"

def distribute_overtime(required_ot, num_working_days):
    """Distributes required OT hours randomly among allowed working days."""
    if num_working_days == 0:
        return []
        
    ot_hours_list = [0] * num_working_days
    hours_distributed = 0
    
    max_attempts = required_ot * 5 
    attempts = 0
    
    while hours_distributed < required_ot and attempts < max_attempts:
        attempts += 1
        day_index = random.randint(0, num_working_days - 1)
        ot_to_add = random.choice([1, 1, 2])
        
        if hours_distributed + ot_to_add > required_ot:
            ot_to_add = required_ot - hours_distributed
            
        if ot_hours_list[day_index] < 2:
           ot_to_add_today = min(ot_to_add, 2 - ot_hours_list[day_index])
           ot_hours_list[day_index] += ot_to_add_today
           hours_distributed += ot_to_add_today
        
        if all(ot >= 2 for ot in ot_hours_list):
            if hours_distributed < required_ot:
                remaining = required_ot - hours_distributed
                for _ in range(remaining):
                    day_index = random.randint(0, num_working_days - 1)
                    ot_hours_list[day_index] += 1
                hours_distributed = sum(ot_hours_list)
            break
            
    return ot_hours_list

def generate_attendance_file(input_df, target_month, target_year, holidays_dict, company_name_input, std_shift, sp_shift=None):
    output = io.BytesIO()
    month_year_str = f"{datetime.date(target_year, target_month, 1).strftime('%B %Y').upper()}"

    title_font = Font(name='Calibri', size=14, bold=True)
    header_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    link_font = Font(name='Calibri', size=11, color="0000FF", underline="single")
    thin_side = Side(border_style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    index_data = []

    def is_special(date_obj):
        """Checks if a given date falls within the special shift date range."""
        if not sp_shift: return False
        return sp_shift["start"] <= date_obj <= sp_shift["end"]
        
    def get_val(row_s, *keys, default=None):
        """Safely fetch a value from the pandas row checking multiple possible column names (Case-Insensitive)."""
        for k in keys:
            if k in row_s:
                return row_s[k]
        return default

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        index_ws = writer.book.create_sheet(title="Index", index=0)
        
        progress_bar = st.progress(0)
        total_emps = len(input_df)
        
        for i, (_, employee) in enumerate(input_df.iterrows()):
            progress_bar.progress((i + 1) / total_emps)

            # --- Robust Data Extraction (Case Insensitive Support) ---
            emp_code = get_val(employee, 'CODE', 'Code', 'code', default='')
            emp_name = get_val(employee, 'NAME', 'Name', 'name', default='')
            s_no = get_val(employee, 'S#', 'S.No', 'S. No', 's#', default='')
            
            req_ot_raw = get_val(employee, 'Overtime Hours', 'OVERTIME HOURS', 'Overtime', default=0)
            req_ot = int(req_ot_raw) if pd.notna(req_ot_raw) and str(req_ot_raw).strip() != '' else 0
            
            abs_raw = get_val(employee, 'ABSENT DAYS', 'Absent Days', 'Absent', default=0)
            try:
                num_absent = int(abs_raw) if pd.notna(abs_raw) and str(abs_raw).strip() != '' else 0
            except ValueError:
                num_absent = 0
            
            # Extract Status (Handling New and Left Employees dynamically)
            emp_status_raw = get_val(employee, 'STATUS', 'Status', 'status', default='')
            if pd.isna(emp_status_raw) or str(emp_status_raw).strip().lower() in ['nan', '']:
                emp_status = ""
            else:
                emp_status = str(emp_status_raw).strip().title()
                
            # Extract Date
            emp_date_raw = get_val(employee, 'DATE', 'Date', 'date', default=pd.NaT)
            # ---------------------------------------------------------
                
            safe_name = str(emp_name).replace(":", "").replace("/", "")
            sheet_name = f"{emp_code}_{safe_name}"[:31]
            ws = writer.book.create_sheet(title=sheet_name)
            
            # Header Data
            header_data = [
                ["Company Name:", company_name_input],
                ["Report Title:", f"ATTENDANCE SHEETS FOR THE MONTH OF {month_year_str}"],
                ["Employee Name:", emp_name],
                ["Employee Code:", emp_code]
            ]
            
            try:
                next_month = datetime.date(target_year, target_month, 28) + datetime.timedelta(days=4)
                last_day_of_month = next_month - datetime.timedelta(days=next_month.day)
                num_days_in_month = last_day_of_month.day
            except ValueError:
                num_days_in_month = 30
                last_day_of_month = datetime.date(target_year, target_month, 30)

            # --- Determine Active Period bounds based on Joining/Leaving status ---
            active_start_date = datetime.date(target_year, target_month, 1)
            active_end_date = last_day_of_month

            if pd.notna(emp_date_raw) and str(emp_date_raw).strip() != "":
                try:
                    parsed_date = pd.to_datetime(emp_date_raw).date()
                    if emp_status == "New":
                        active_start_date = max(active_start_date, parsed_date)
                    elif emp_status == "Left":
                        active_end_date = min(active_end_date, parsed_date)
                except Exception:
                    pass # Ignore if date format is strictly invalid
                
            working_days_in_month = []
            full_month_data = []
            sundays = 0
            holidays_found = 0
            
            for day_num in range(1, num_days_in_month + 1):
                current_date = datetime.date(target_year, target_month, day_num)
                
                # Only check holidays and sundays if the employee is currently active
                if active_start_date <= current_date <= active_end_date:
                    is_sunday = current_date.weekday() == 6
                    is_holiday = current_date in holidays_dict
                    
                    if is_sunday:
                        sundays += 1
                    elif is_holiday:
                        holidays_found += 1
                    else:
                        working_days_in_month.append(current_date)
            
            num_working_days = len(working_days_in_month)
            absent_days = set()
            
            # Make sure we don't assign more absent days than the employee actually worked
            actual_absent = min(num_absent, num_working_days) 
            if actual_absent > 0:
                absent_days = set(random.sample(working_days_in_month, actual_absent))
                
            index_data.append({
                "S. No": s_no,
                "CODE": emp_code,
                "Name": emp_name,
                "SheetName": sheet_name,
                "Absent": actual_absent,
                "OT Hours": req_ot,
                "Status": emp_status
            })
            
            # Divide working days into standard and special
            working_days_with_attendance = [day for day in working_days_in_month if day not in absent_days]
            standard_working_days = [day for day in working_days_with_attendance if not is_special(day)]
            
            # Distribute OT *only* among standard working days
            ot_schedule = distribute_overtime(req_ot, len(standard_working_days))
            
            std_work_counter = 0
            sp_work_counter = 0
            total_std_hours = 0
            
            for day_num in range(1, num_days_in_month + 1):
                current_date = datetime.date(target_year, target_month, day_num)
                date_str = current_date.strftime("%d-%b-%y")
                
                row = [date_str, std_shift['name'], "", "", "", ""]

                # Process dates before joining or after leaving first
                if current_date < active_start_date:
                    row[1] = "-"
                    row[5] = "Not Joined"
                elif current_date > active_end_date:
                    row[1] = "-"
                    row[5] = "Left"
                else:
                    is_sunday = current_date.weekday() == 6
                    holiday_name = holidays_dict.get(current_date)
                    
                    if is_sunday:
                        row[5] = "SUNDAY"
                    elif holiday_name:
                        row[5] = holiday_name
                    elif current_date in working_days_with_attendance:
                        if is_special(current_date):
                            # Special Shift Day Logic
                            row[1] = sp_shift['name']
                            row[2] = create_natural_time(target_year, target_month, 9, True)
                            row[3] = create_natural_time(target_year, target_month, sp_shift['out_hour'], False)
                            row[4] = "" # NO OT FOR SPECIAL SHIFT
                            row[5] = "On Time"
                            total_std_hours += sp_shift['hours']
                            sp_work_counter += 1
                        else:
                            # Standard Shift Day Logic
                            ot_hours = ot_schedule[std_work_counter]
                            row[1] = std_shift['name']
                            row[2] = create_natural_time(target_year, target_month, 9, True)
                            row[3] = create_natural_time(target_year, target_month, std_shift['out_hour'] + ot_hours, False)
                            row[4] = ot_hours if ot_hours > 0 else ""
                            row[5] = "On Time"
                            total_std_hours += std_shift['hours']
                            std_work_counter += 1
                    elif current_date in working_days_in_month:
                        row[5] = "Absent"
                
                full_month_data.append(row)
                
            # Footing Logic
            total_present_days = std_work_counter + sp_work_counter
            total_ot_hours = sum(ot_schedule)
            total_payable_hours = total_std_hours + total_ot_hours
            
            shift_breakdown_str = f"({std_work_counter} Std. Days x {std_shift['hours']}h)"
            if sp_work_counter > 0:
                shift_breakdown_str += f" + ({sp_work_counter} Spc. Days x {sp_shift['hours']}h)"
            
            footing_data = [
                ["SUMMARY:", ""],
                ["Total Days in Month", num_days_in_month],
                ["Sundays (Active)", sundays],
                ["Gazetted Holidays (Active)", holidays_found],
                ["Total Present Days", total_present_days],
                ["Absent", actual_absent],
                ["Over Time Hrs.", total_ot_hours],
                [],
                ["Total Standard Hours", total_std_hours, shift_breakdown_str],
                ["Total OT Hours", total_ot_hours, f"(Sum of OT HRS)"],
                ["Total Payable Hours", total_payable_hours]
            ]
            
            # --- WRITING & FORMATTING ---
            
            # 1. Header Formatting
            ws.cell(row=1, column=1, value="Company Name:").font = title_font
            ws.cell(row=1, column=2, value=company_name_input).font = title_font 
            ws.merge_cells('B1:F1') 

            for r_idx, row_val in enumerate(header_data[1:], 2):
                ws.cell(row=r_idx, column=1, value=row_val[0]).font = header_font
                ws.cell(row=r_idx, column=2, value=row_val[1]).font = normal_font
                ws.merge_cells(f'B{r_idx}:F{r_idx}') 

            # 2. Data Table
            data_start_row = 6
            table_headers = ["DATE", "SHIFT G", "TIME IN", "TIME OUT", "OT HRS", "REMARKS"]
            
            for c_idx, val in enumerate(table_headers, 1):
                cell = ws.cell(row=data_start_row, column=c_idx, value=val)
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align
                
            for r_idx, row_val in enumerate(full_month_data, data_start_row + 1):
                for c_idx, val in enumerate(row_val, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = center_align
                    if c_idx == 5: cell.alignment = right_align

            # 3. Footing
            footing_start_row = data_start_row + len(full_month_data) + 2
            ws.cell(row=footing_start_row, column=1, value="SUMMARY:").font = header_font
            ws.cell(row=footing_start_row, column=1).border = thin_border
            
            for r_idx, row_val in enumerate(footing_data[1:], footing_start_row + 1):
                for c_idx, val in enumerate(row_val, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = normal_font
                    cell.border = thin_border
                    if c_idx == 1: cell.font = header_font
                    if c_idx > 1: cell.alignment = right_align

            # 4. Dimensions & Print
            widths = [15, 15, 12, 12, 10, 20]
            for i, w in enumerate(widths):
                ws.column_dimensions[get_column_letter(i+1)].width = w
                
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1

        # --- Index Sheet Logic ---
        index_headers = ["S. No", "CODE", "Name", "Absent", "OT Hours", "Status"]
        for c_idx, val in enumerate(index_headers, 1):
            cell = index_ws.cell(row=1, column=c_idx, value=val)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
            
        for r_idx, data in enumerate(index_data, 2):
            c1 = index_ws.cell(row=r_idx, column=1, value=data['S. No'])
            c1.font = normal_font; c1.border = thin_border; c1.alignment = center_align
            c2 = index_ws.cell(row=r_idx, column=2, value=data['CODE'])
            c2.font = normal_font; c2.border = thin_border; c2.alignment = center_align
            c3 = index_ws.cell(row=r_idx, column=3)
            c3.value = f'=HYPERLINK("#\'{data["SheetName"]}\'!A1", "{data["Name"]}")'
            c3.font = link_font; c3.border = thin_border
            c4 = index_ws.cell(row=r_idx, column=4, value=data['Absent'])
            c4.font = normal_font; c4.border = thin_border; c4.alignment = center_align
            c5 = index_ws.cell(row=r_idx, column=5, value=data['OT Hours'])
            c5.font = normal_font; c5.border = thin_border; c5.alignment = center_align
            c6 = index_ws.cell(row=r_idx, column=6, value=data['Status'])
            c6.font = normal_font; c6.border = thin_border; c6.alignment = center_align
            
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])

    return output

# --- B. INVOICE HELPERS ---
def num_to_words(n):
    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine', 'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']

    def convert(n):
        if n < 20: return ones[n]
        if n < 100: return tens[n // 10] + ('' if n % 10 == 0 else ' ' + ones[n % 10])
        if n < 1000: return ones[n // 100] + ' Hundred' + ('' if n % 100 == 0 else ' and ' + convert(n % 100))
        if n < 1000000: return convert(n // 1000) + ' Thousand' + ('' if n % 1000 == 0 else ' ' + convert(n % 1000))
        if n < 1000000000: return convert(n // 1000000) + ' Million' + ('' if n % 1000000 == 0 else ' ' + convert(n % 1000000))
        return 'Number too large'

    if n == 0: return 'Zero'
    
    num_str = f"{n:.2f}"
    integer_part, decimal_part = num_str.split('.')
    words = convert(int(integer_part))
    if int(decimal_part) > 0:
        words += " and " + convert(int(decimal_part)) + " Paisa"
    return words + " Only"

def generate_html_invoice(input_df, header_info, tax_rate):
    grouped = input_df.groupby('DC No.')
    all_invoices_html = ""
    
    for dc_no, group in grouped:
        header_row = group.iloc[0]
        customer_name = header_row.get('Customer Name', '')
        bill_address = header_row.get('Bill To Address', '')
        customer_ntn = header_row.get('Customer NTN', '')
        invoice_no = header_row.get('Invoice No.', '')
        
        raw_date = header_row.get('Invoice Date', '')
        try:
            invoice_date = pd.to_datetime(raw_date).strftime('%d-%b-%Y')
        except:
            invoice_date = str(raw_date)
            
        payment_terms = header_row.get('Credit Terms', 'Cash')
        
        sub_total = group['Total Value (PKR)'].sum()
        tax_amount = sub_total * (tax_rate / 100)
        grand_total = sub_total + tax_amount
        amount_in_words = num_to_words(grand_total)
        
        rows_html = ""
        for idx, row in group.iterrows():
            u_price = f"{row['Unit Price (PKR)']:,.2f}"
            t_value = f"{row['Total Value (PKR)']:,.2f}"
            
            rows_html += f"""
            <tr class="bg-white">
                <td class="p-1 text-center">{idx + 1}</td>
                <td class="p-1">{row['H.S Code']}</td>
                <td class="p-1 wrap-text">{row['Item Description']}</td>
                <td class="p-1">Weaving</td> 
                <td class="p-1">JOB-{random.randint(1000,9999)}</td>
                <td class="p-1">{row['DC No.']}</td>
                <td class="p-1">{row['UOM']}</td>
                <td class="p-1 text-center">{row['Qty']}</td>
                <td class="p-1 text-right">{u_price}</td>
                <td class="p-1 text-right">{t_value}</td>
            </tr>
            """
            
        for _ in range(max(0, 8 - len(group))):
             rows_html += '<tr class="bg-white"><td class="p-2 text-center">&nbsp;</td><td></td><td class="wrap-text"></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>'

        invoice_html = f"""
        <div class="printable-container max-w-6xl mx-auto bg-white p-6 md:p-8 mb-8" style="page-break-after: always;">
            <header class="flex justify-between items-start pb-4">
                <div>
                    <h1 class="text-2xl md:text-3xl font-bold text-gray-800">{header_info['company_name']}</h1>
                    <p class="text-sm text-gray-500">{header_info['address']}</p>
                    <p class="text-sm text-gray-500">Phones: {header_info['phone']}</p>
                    <p class="text-sm text-gray-500">E-mail: {header_info['email']} | Website: {header_info['web']}</p>
                    <p class="text-sm text-gray-500 font-semibold mt-1">NTN: {header_info['ntn']}</p>
                </div>
                <div class="text-right">
                    <h2 class="text-2xl md:text-3xl font-semibold text-gray-700">SALES TAX INVOICE</h2>
                    <div class="mt-2 grid grid-cols-2 gap-2 text-left">
                        <label class="block text-xs font-medium text-gray-500 p-1">Invoice No.</label>
                        <input type="text" value="{invoice_no}" class="block w-full p-1 border-2 border-black rounded-md shadow-sm text-sm text-black font-medium" readonly>
                        
                        <label class="block text-xs font-medium text-gray-500 p-1">Invoice Date</label>
                        <input type="text" value="{invoice_date}" class="block w-full p-1 border-2 border-black rounded-md shadow-sm text-sm text-black font-medium">
                        
                        <label class="block text-xs font-medium text-gray-500 p-1">Payment Terms</label>
                        <input type="text" value="{payment_terms}" class="block w-full p-1 border-2 border-black rounded-md shadow-sm text-sm text-black font-medium">

                        <label class="block text-xs font-medium text-gray-500 p-1">Customer PO</label>
                        <input type="text" value="PO-REF-XX" class="block w-full p-1 border-2 border-black rounded-md shadow-sm text-sm text-black font-medium">
                    </div>
                </div>
            </header>

            <div class="main-content">
                <section class="grid grid-cols-2 gap-6 mt-6 section-spacing">
                    <div class="border-2 border-black rounded-md p-3">
                        <h3 class="text-sm font-semibold text-white mb-2 bg-gray-700 p-1 -m-3 border-b border-black dark-bg print-header">BILL TO</h3>
                        <div class="mt-3">
                            <label class="block text-xs font-medium text-black font-bold">Customer Name</label>
                            <input type="text" value="{customer_name}" class="mt-1 block w-full p-2 border-2 border-black rounded-md shadow-sm text-sm text-black dark-border font-medium">
                        </div>
                        <div class="mt-2">
                            <label class="block text-xs font-medium text-black font-bold">Address</label>
                            <textarea class="mt-1 block w-full p-2 border-2 border-black rounded-md shadow-sm text-sm text-black dark-border font-medium" rows="2">{bill_address}</textarea>
                        </div>
                        <div class="mt-2 grid grid-cols-2 gap-2">
                             <div>
                                <label class="block text-xs font-medium text-black font-bold">NTN</label>
                                <input type="text" value="{customer_ntn}" class="mt-1 block w-full p-2 border-2 border-black rounded-md shadow-sm text-sm text-black dark-border font-medium">
                             </div>
                             <div>
                                <label class="block text-xs font-medium text-black font-bold">STRN</label>
                                <input type="text" value="" class="mt-1 block w-full p-2 border-2 border-black rounded-md shadow-sm text-sm text-black dark-border font-medium">
                             </div>
                        </div>
                    </div>
                    <div class="border-2 border-black rounded-md p-3">
                        <h3 class="text-sm font-semibold text-white mb-2 bg-gray-700 p-1 -m-3 border-b border-black dark-bg print-header">SHIP TO</h3>
                        <div class="mt-3">
                            <p class="text-sm font-semibold text-black">{customer_name}</p>
                            <p class="text-sm text-black">{bill_address}</p>
                        </div>
                    </div>
                </section>
    
                <section class="mt-6 table-container section-spacing">
                    <h3 class="text-lg font-semibold text-gray-700 mb-2">Item Details</h3>
                    <table class="w-full text-sm text-left text-gray-500 printable-table">
                        <thead class="text-xs text-white uppercase bg-gray-700 dark-bg print-header" style="-webkit-print-color-adjust: exact;">
                            <tr>
                                <th class="p-2 text-center" style="width: 3%;">Sr.</th>
                                <th class="p-2" style="width: 8%;">H.S Code</th>
                                <th class="p-2 wrap-text" style="width: 25%;">Item Description</th>
                                <th class="p-2" style="width: 10%;">Cost Center</th>
                                <th class="p-2" style="width: 10%;">Job No.</th>
                                <th class="p-2" style="width: 10%;">DC No.</th>
                                <th class="p-2" style="width: 4%;">UOM</th>
                                <th class="p-2 text-center" style="width: 5%;">Qty</th>
                                <th class="p-2 text-right" style="width: 10%;">Unit Price</th>
                                <th class="p-2 text-right" style="width: 15%;">Total Value</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </section>
                
                <section class="grid grid-cols-2 gap-6 mt-6 section-spacing">
                    <div>
                        <label class="block text-sm font-medium text-black font-bold">Amount in Words (PKR)</label>
                        <textarea class="mt-1 block w-full p-2 border-2 border-black rounded-md shadow-sm text-sm dark-border text-black" rows="2" readonly>{amount_in_words}</textarea>
                    </div>
                    <div class="space-y-2">
                        <div class="flex justify-between items-center bg-gray-700 text-white p-2 rounded-md border-2 border-black dark-bg print-total-box">
                            <span class="text-sm font-bold text-white">Sub-Total:</span>
                            <span class="text-sm font-bold text-white">{sub_total:,.2f}</span>
                        </div>
                        <div class="flex justify-between items-center p-2">
                            <div class="text-sm font-bold text-black">
                                Sales Tax ({tax_rate}%):
                            </div>
                            <span class="text-sm font-bold text-black">{tax_amount:,.2f}</span>
                        </div>
                        <div class="flex justify-between items-center bg-gray-700 text-white p-3 rounded-md border-2 border-black dark-bg print-total-box">
                            <span class="text-base font-bold text-white">Grand Total:</span>
                            <span class="text-base font-bold text-white">{grand_total:,.2f}</span>
                        </div>
                    </div>
                </section>
            </div>

            <footer class="mt-8">
                <div class="grid grid-cols-2 gap-8">
                    <div></div>
                    <div class="text-center">
                        <p class="signature-line pt-2 text-sm font-semibold text-gray-700">For {header_info['company_name']}</p>
                        <p class="text-xs text-gray-500">(Authorized Signatory)</p>
                    </div>
                </div>
            </footer>
        </div>
        """
        all_invoices_html += invoice_html

    full_html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Sales Tax Invoices</title>
        <script src="https://cdn.tailwindcss.com"></script>
        <style>
            body {{ background-color: #f9fafb; font-family: Calibri, sans-serif; }}
            @media print {{
                @page {{ size: A4 portrait; margin: 0.1cm; margin-bottom: 0.5cm; }}
                html, body {{ background-color: #fff; font-size: 9pt; }}
                .no-print {{ display: none; }}
                input, textarea, select {{ border: none !important; resize: none; }}
                .printable-table th {{ background-color: #374151 !important; color: #ffffff !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
                .print-header {{ background-color: #374151 !important; color: #ffffff !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
                .print-total-box {{ background-color: #374151 !important; color: #ffffff !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
                .text-black {{ color: #000000 !important; }}
                .text-gray-500, .text-gray-600, .text-gray-700, .text-gray-800 {{ color: #000000 !important; }}
                .border-black {{ border-color: #000000 !important; border-width: 2px !important; border-style: solid !important; }}
                .border-2 {{ border-width: 2px !important; }}
            }}
            .signature-line {{ border-top: 1px solid #4A5568; margin-top: 2.5rem; }}
            .printable-table, .printable-table th, .printable-table td {{ border: 1px solid #000000 !important; border-collapse: collapse; }}
        </style>
    </head>
    <body class="p-4 md:p-8">
        {all_invoices_html}
        <div class="fixed bottom-4 right-4 no-print">
            <button onclick="window.print()" class="px-6 py-3 bg-blue-600 text-white font-bold rounded-full shadow-lg hover:bg-blue-700 transition">
                🖨️ Print Invoices
            </button>
        </div>
    </body>
    </html>
    """
    return full_html

def generate_excel_invoice(input_df, header_info, tax_rate):
    output = io.BytesIO()
    grouped = input_df.groupby('DC No.')
    header_font = Font(name='Calibri', size=14, bold=True)
    sub_header_font = Font(name='Calibri', size=10)
    table_header_font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
    fill_dark = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        ws = writer.book.create_sheet("Invoices")
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        
        current_row = 1
        for dc_no, group in grouped:
            header_row = group.iloc[0]
            invoice_no = header_row.get('Invoice No.', '')
            raw_date = header_row.get('Invoice Date', '')
            invoice_date = raw_date if isinstance(raw_date, str) else raw_date.strftime('%d-%b-%Y')
            
            ws.cell(row=current_row, column=1, value=header_info['company_name']).font = header_font
            ws.cell(row=current_row, column=8, value="SALES TAX INVOICE").font = header_font
            current_row += 1
            ws.cell(row=current_row, column=1, value=header_info['address']).font = sub_header_font
            ws.cell(row=current_row, column=8, value=f"Invoice No: {invoice_no}").font = sub_header_font
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"Phone: {header_info['phone']}").font = sub_header_font
            ws.cell(row=current_row, column=8, value=f"Date: {invoice_date}").font = sub_header_font
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"NTN: {header_info['ntn']}").font = sub_header_font
            current_row += 2 
            
            ws.cell(row=current_row, column=1, value="BILL TO").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=header_row.get('Customer Name', ''))
            current_row += 1
            ws.cell(row=current_row, column=1, value="Address").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=header_row.get('Bill To Address', ''))
            current_row += 1
            ws.cell(row=current_row, column=1, value="NTN").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=header_row.get('Customer NTN', ''))
            current_row += 2
            
            headers = ["Sr.", "H.S Code", "Description", "Cost Center", "Job No", "DC No", "UOM", "Qty", "Unit Price", "Total"]
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(row=current_row, column=col_idx, value=h)
                c.font = table_header_font
                c.fill = fill_dark
                c.alignment = Alignment(horizontal='center')
            current_row += 1
            
            sub_total = 0
            for idx, row in group.iterrows():
                ws.cell(row=current_row, column=1, value=idx+1).border = thin_border
                ws.cell(row=current_row, column=2, value=row['H.S Code']).border = thin_border
                ws.cell(row=current_row, column=3, value=row['Item Description']).border = thin_border
                ws.cell(row=current_row, column=4, value="Weaving").border = thin_border
                ws.cell(row=current_row, column=5, value="JOB-XXXX").border = thin_border
                ws.cell(row=current_row, column=6, value=row['DC No.']).border = thin_border
                ws.cell(row=current_row, column=7, value=row['UOM']).border = thin_border
                ws.cell(row=current_row, column=8, value=row['Qty']).border = thin_border
                ws.cell(row=current_row, column=9, value=row['Unit Price (PKR)']).border = thin_border
                total_val = row['Total Value (PKR)']
                ws.cell(row=current_row, column=10, value=total_val).border = thin_border
                sub_total += total_val
                current_row += 1
                
            tax_amount = sub_total * (tax_rate / 100)
            grand_total = sub_total + tax_amount
            
            current_row += 1
            ws.cell(row=current_row, column=9, value="Sub-Total").font = Font(bold=True)
            ws.cell(row=current_row, column=10, value=sub_total).font = Font(bold=True)
            current_row += 1
            ws.cell(row=current_row, column=9, value=f"GST ({tax_rate}%)").font = Font(bold=True)
            ws.cell(row=current_row, column=10, value=tax_amount).font = Font(bold=True)
            current_row += 1
            ws.cell(row=current_row, column=9, value="Grand Total").font = Font(bold=True)
            ws.cell(row=current_row, column=10, value=grand_total).font = Font(bold=True)
            
            current_row += 1
            ws.cell(row=current_row, column=1, value="Amount in Words: " + num_to_words(grand_total)).font = Font(italic=True)
            current_row += 4
            
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])
    return output

# --- C. BANK CONVERTER HELPERS ---
def to_float(x):
    if not x: return 0.0
    s = str(x).strip().replace(',', '').replace('(', '-').replace(')', '')
    s = re.sub(r'[^0-9.-]', '', s) 
    try:
        return float(s)
    except:
        return 0.0

def extract_text_from_pdf(pdf_file):
    text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            content = page.extract_text()
            if content:
                text += content + "\n"
    return text

def parse_bank_statement(raw_text):
    lines = raw_text.split('\n')
    date_pattern = r'^(\d{2}/\d{2}/\d{4})'
    
    transactions = []
    current_row = None
    running_balance = 0.0
    account_active = False

    for line in lines:
        line = line.strip()
        if not line: continue

        if "Opening Balance" in line:
            parts = line.split()
            balance_val = to_float(parts[-1])
            running_balance = balance_val
            account_active = True
            
            transactions.append({
                "Posting Date": parts[0] if re.match(r'\d', parts[0]) else "",
                "Value Date": "",
                "Instrument/Doc No": "",
                "Details": "--- Opening Balance ---",
                "Debit": 0.0,
                "Credit": 0.0,
                "Balance": balance_val
            })
            continue

        match = re.match(date_pattern, line)
        if match and account_active:
            if current_row: 
                transactions.append(current_row)
            
            parts = line.split()
            if len(parts) < 2: continue 
            
            post_date = parts[0]
            
            instrument = ""
            details_start_idx = 1
            if len(parts) > 2 and parts[1].isdigit() and 6 <= len(parts[1]) <= 12:
                instrument = parts[1]
                details_start_idx = 2
            elif len(parts) > 3 and parts[2].isdigit() and 6 <= len(parts[2]) <= 12:
                instrument = parts[2]
                details_start_idx = 3

            try:
                row_balance = to_float(parts[-1])
                numeric_candidates = []
                for p in reversed(parts[:-1]):
                    if re.search(r'\d', p) and ('.' in p or ',' in p or p.isdigit()):
                        numeric_candidates.append(p)
                    else:
                        break
                
                diff = round(row_balance - running_balance, 2)
                debit = abs(diff) if diff < 0 else 0.0
                credit = diff if diff > 0 else 0.0
                
                details_end_idx = len(parts) - 1 - len(numeric_candidates)
                details = " ".join(parts[details_start_idx:details_end_idx])

                current_row = {
                    "Posting Date": post_date,
                    "Value Date": post_date, 
                    "Instrument/Doc No": instrument,
                    "Details": details,
                    "Debit": debit,
                    "Credit": credit,
                    "Balance": row_balance
                }
                running_balance = row_balance
            except Exception:
                continue
        else:
            stop_keywords = ["Carried Forward", "Brought Forward", "Page", "Produced On", "TOTALS", "Closing Balance"]
            if current_row and not any(k in line for k in stop_keywords):
                if not re.match(r'^[\d,.\s\-:><]+$', line) or len(line) > 15:
                    current_row["Details"] += " " + line

    if current_row: 
        transactions.append(current_row)
    return transactions

def generate_bank_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Extracted Data')
        workbook = writer.book
        worksheet = writer.sheets['Extracted Data']
        
        header_fmt = workbook.add_format({
            'bold': True, 'font_color': 'white', 'bg_color': '#003366', 
            'border': 1, 'align': 'center', 'valign': 'vcenter'
        })
        num_fmt = workbook.add_format({'num_format': '#,##0.00', 'border': 1, 'valign': 'top', 'font_name': 'Arial', 'font_size': 10})
        date_fmt = workbook.add_format({'num_format': 'dd/mm/yyyy', 'align': 'center', 'border': 1, 'valign': 'top', 'font_name': 'Arial', 'font_size': 10})
        text_fmt = workbook.add_format({'text_wrap': True, 'border': 1, 'valign': 'top', 'font_name': 'Arial', 'font_size': 10})
        ob_num = workbook.add_format({'num_format': '#,##0.00', 'border': 1, 'bg_color': '#FFFFCC', 'bold': True, 'valign': 'top'})
        ob_text = workbook.add_format({'text_wrap': True, 'border': 1, 'bg_color': '#FFFFCC', 'bold': True, 'valign': 'top'})
        ob_date = workbook.add_format({'align': 'center', 'border': 1, 'bg_color': '#FFFFCC', 'bold': True, 'valign': 'top'})
        
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_fmt)
            
        for row_idx, row_data in enumerate(df.values):
            details_str = str(row_data[3])
            is_ob = "Opening Balance" in details_str
            d_fmt = ob_date if is_ob else date_fmt
            t_fmt = ob_text if is_ob else text_fmt
            n_fmt = ob_num  if is_ob else num_fmt
            
            worksheet.write(row_idx + 1, 0, row_data[0], d_fmt)
            worksheet.write(row_idx + 1, 1, row_data[1], d_fmt)
            worksheet.write(row_idx + 1, 2, row_data[2], t_fmt)
            worksheet.write(row_idx + 1, 3, row_data[3], t_fmt)
            worksheet.write(row_idx + 1, 4, row_data[4], n_fmt)
            worksheet.write(row_idx + 1, 5, row_data[5], n_fmt)
            worksheet.write(row_idx + 1, 6, row_data[6], n_fmt)
            
        worksheet.set_column('A:B', 14)
        worksheet.set_column('C:C', 18)
        worksheet.set_column('D:D', 65)
        worksheet.set_column('E:G', 18)
    return output.getvalue()

# --- D. TAX CALCULATION HELPER ---
def calculate_fbr_tax(monthly_gross_salary):
    annual_income = monthly_gross_salary * 12
    annual_tax = 0

    if annual_income <= 600000:
        annual_tax = 0
    elif annual_income <= 1200000:
        annual_tax = (annual_income - 600000) * 0.01
    elif annual_income <= 2200000:
        annual_tax = 6000 + (annual_income - 1200000) * 0.11
    elif annual_income <= 3200000:
        annual_tax = 116000 + (annual_income - 2200000) * 0.23
    elif annual_income <= 4100000:
        annual_tax = 346000 + (annual_income - 3200000) * 0.30
    else:
        annual_tax = 616000 + (annual_income - 4100000) * 0.35

    monthly_tax = annual_tax / 12
    return annual_income, annual_tax, monthly_tax

# ==========================================
# 3. SIDEBAR
# ==========================================
with st.sidebar:
    try:
        logo = PILImage.open("nfp_office.jpg") 
        st.image(logo, use_container_width=True)
    except:
        try:
            logo = PILImage.open("Nazeer Fin Pro - NFP.jpg")
            st.image(logo, use_container_width=True)
        except:
            st.header(f"✨ {BRAND_SHORT}")
        
    st.markdown(f"## **{BRAND_NAME}**")
    st.caption("Professional Finance Consultancy")
    st.write("---")
    
    st.write("**🎥 App Tutorial**")
    if os.path.exists("Tutorial.mp4"):
        st.video("Tutorial.mp4")
    else:
        st.info("Tutorial video not available.")
    st.write("---")
    
    col_about_label, col_about_img = st.columns([3, 1])
    with col_about_label:
        st.write("**About Nazeer Ahmed Khan**")
    with col_about_img:
        try:
            about_img = PILImage.open("about_nazeer.jpg")
            st.image(about_img, use_container_width=True)
        except:
            pass
            
    with st.expander("View Details"):
        st.write("""
        **Nazeer Ahmed Khan** is the founder of NazeerFinPro (NFP). 
        
        A seasoned Finance Freelance Consultant specializing in:
        - Financial Analysis & Reporting
        - Automation & Data Cleaning (Python/Excel)
        - Business Intelligence Dashboards
        - Corporate Training
        
        *\"Turning chaotic data into clear, actionable financial insights.\"*
        """)
    st.write("---")

# ==========================================
# 4. MAIN CONTENT
# ==========================================

col_title, col_logo = st.columns([5, 1])

with col_title:
    st.markdown(f"<div class='brand-header'>{BRAND_NAME} Tool Suite</div>", unsafe_allow_html=True)
    st.markdown("<div class='brand-sub'>Advanced Financial Solutions & Automation</div>", unsafe_allow_html=True)

with col_logo:
    try:
        header_logo = PILImage.open("logo.jpg")
        st.image(header_logo, width=100) 
    except:
        pass 

# TABS
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["🏢 Attendance Generator", "🧾 Invoice Maker", "🏦 Bank Converter", "🧮 Payroll Calculator", "📝 Consultancy Blog", "📞 Contact NFP"])

# --- TAB 1: ATTENDANCE ---
with tab1:
    st.subheader("Auto-Generate Attendance Sheets")
    st.info("Upload your employee data file (`data.xlsx`) to generate payroll-ready Excel sheets with natural time variations.")
    
    if os.path.exists("data.xlsx"):
        with open("data.xlsx", "rb") as template_file:
            st.download_button(
                label="📥 Download Sample Template",
                data=template_file,
                file_name="data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("⚠️ 'data.xlsx' template not found in repository.")
    
    with st.expander("⚙️ Customize Report Header & Settings", expanded=True):
        col_gen_1, col_gen_2 = st.columns(2)
        with col_gen_1:
            company_name = st.text_input("Company Name", value="ABC COMPANY", key="att_company_name")
            target_date = st.date_input("Select Month & Year", datetime.date(2026, 2, 1), key="att_target_date")
            selected_month = target_date.month
            selected_year = target_date.year
        
        with col_gen_2:
            st.write("**Gazetted Holidays**")
            if 'holidays' not in st.session_state:
                st.session_state.holidays = []
            
            col_h1, col_h2 = st.columns(2)
            with col_h1:
                default_date_val = datetime.date(selected_year, selected_month, 1)
                h_date = st.date_input("Holiday Date", value=default_date_val, label_visibility="collapsed")
            with col_h2:
                h_name = st.text_input("Holiday Name", placeholder="Name", label_visibility="collapsed")
                
            col_b1, col_b2 = st.columns(2)
            with col_b1:
                if st.button("Add Holiday", key="add_hol_btn"):
                    if h_name:
                        st.session_state.holidays.append({"date": h_date, "name": h_name})
                        st.success(f"Added: {h_name}")
                    else:
                        st.error("Enter Name")
            with col_b2:
                if st.button("Clear Holidays", key="clear_hol_btn"):
                    st.session_state.holidays = []
                    st.rerun()

            holidays_dict = {}
            current_month_holidays = []
            if st.session_state.holidays:
                for h in st.session_state.holidays:
                    if h['date'].year == selected_year and h['date'].month == selected_month:
                        current_month_holidays.append(h)
                        if h['date'] in holidays_dict:
                            holidays_dict[h['date']] += f" / {h['name']}"
                        else:
                            holidays_dict[h['date']] = h['name']
            
            if current_month_holidays:
                st.caption("Active Holidays:")
                for h in current_month_holidays:
                    st.caption(f"- {h['date'].strftime('%d-%b')}: {h['name']}")
            else:
                st.caption("No holidays for this month.")

    # NEW: SHIFT & OVERTIME SETTINGS EXPANDER
    with st.expander("⏱️ Shift & Overtime Settings (Standard & Special)", expanded=False):
        st.write("**Standard Working Days Settings**")
        col_s1, col_s2, col_s3 = st.columns(3)
        with col_s1: std_shift_name = st.text_input("Standard Shift Name", "(0900:1800)")
        with col_s2: std_hours = st.number_input("Standard Hours/Day", value=9, min_value=1)
        with col_s3: std_out_hour = st.number_input("Standard Checkout Hour (24h)", value=18, min_value=1, max_value=23)
        
        std_shift_config = {"name": std_shift_name, "hours": std_hours, "out_hour": std_out_hour}

        st.write("---")
        use_special_shift = st.checkbox("✅ Apply Special Shift (e.g., Ramadan Timings)", value=False)
        special_shift_config = None
        
        if use_special_shift:
            st.info("During these dates, Overtime will NOT be calculated. OT will only be distributed among Standard days.")
            col_sp1, col_sp2 = st.columns(2)
            with col_sp1:
                # Setup default for Feb 19 - Feb 28 as requested
                sp_start = st.date_input("Special Shift Start Date", datetime.date(selected_year, selected_month, 19) if selected_month == 2 else datetime.date(selected_year, selected_month, 1))
                sp_end = st.date_input("Special Shift End Date", datetime.date(selected_year, selected_month, 28) if selected_month == 2 else datetime.date(selected_year, selected_month, 15))
            with col_sp2:
                sp_shift_name = st.text_input("Special Shift Name", "(0900:1600)")
                sp_hours = st.number_input("Special Hours/Day", value=7, min_value=1)
                sp_out_hour = st.number_input("Special Checkout Hour (24h)", value=16, min_value=1, max_value=23)

            special_shift_config = {
                "start": sp_start,
                "end": sp_end,
                "name": sp_shift_name,
                "hours": sp_hours,
                "out_hour": sp_out_hour
            }

    uploaded_file = st.file_uploader("Upload Input File", type=['xlsx'])

    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            st.success("File loaded!")
            with st.expander("View Input Data"):
                st.dataframe(df.head())
                
            if st.button("🚀 Generate & Download Report", type="primary"):
                with st.spinner("Processing data..."):
                    excel_data = generate_attendance_file(df, selected_month, selected_year, holidays_dict, company_name, std_shift_config, special_shift_config)
                    st.success("Done! Your file is ready.")
                    file_name = f"NFP_Attendance_{target_date.strftime('%B_%Y')}.xlsx"
                    st.download_button(
                        label="📥 Download Excel File",
                        data=excel_data.getvalue(),
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        except Exception as e:
            st.error(f"Error: {e}")

# --- TAB 2: INVOICE MAKER ---
with tab2:
    st.subheader("🧾 Invoice Maker")
    st.info("Upload your Sales Register (`sales_register.xlsx`) to generate bulk GST invoices ready for printing.")
    
    if os.path.exists("sales_register.xlsx"):
        with open("sales_register.xlsx", "rb") as template_file:
            st.download_button(
                label="📥 Download Sales Register Template",
                data=template_file,
                file_name="sales_register.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("⚠️ 'sales_register.xlsx' template not found in repository.")

    with st.expander("⚙️ Customize Company Header & Tax", expanded=True):
        col_inv_1, col_inv_2 = st.columns(2)
        with col_inv_1:
            inv_company_name = st.text_input("Company Name", value="NazeerFinPro-NFP")
            inv_address = st.text_area("Company Address", value="Plot No. 123, S.I.T.E, Karachi, Pakistan.")
            inv_phone = st.text_input("Contact No.", value="00923333126614")
        with col_inv_2:
            inv_email = st.text_input("Email", value="nfp@gmail.com")
            inv_web = st.text_input("Web Address", value="www.nfp.com")
            inv_ntn = st.text_input("Company NTN", value="N123456-7")
            inv_tax_rate = st.number_input("Sales Tax Rate (%)", value=18.0, step=1.0)
            
    header_info = {
        "company_name": inv_company_name,
        "address": inv_address,
        "phone": inv_phone,
        "email": inv_email,
        "web": inv_web,
        "ntn": inv_ntn
    }
    
    invoice_file = st.file_uploader("Upload Sales Register", type=['xlsx'], key="invoice_uploader")
    
    if invoice_file is not None:
        try:
            inv_df = pd.read_excel(invoice_file)
            st.success("Sales Register Loaded!")
            with st.expander("Preview Sales Data"):
                st.dataframe(inv_df.head())
            if st.button("🖨️ Generate Printable Invoices", type="primary"):
                with st.spinner("Generating Invoices..."):
                    html_content = generate_html_invoice(inv_df, header_info, inv_tax_rate)
                    excel_inv_data = generate_excel_invoice(inv_df, header_info, inv_tax_rate)
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        st.download_button(
                            label="📥 Download Invoice (HTML)",
                            data=html_content,
                            file_name="GST_Invoices_Printable.html",
                            mime="text/html"
                        )
                    with col_d2:
                        st.download_button(
                            label="📥 Download Excel Invoices",
                            data=excel_inv_data.getvalue(),
                            file_name="GST_Invoices.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        except Exception as e:
            st.error(f"Error processing file: {e}")

# --- TAB 3: BANK CONVERTER ---
with tab3:
    st.subheader("🏦 Bank Statement Converter Pro")
    st.markdown("Automated PDF to Excel Extraction for **Bank AL Habib**.")
    st.info("Direct PDF upload is the most accurate method. The app will automatically calculate Debit/Credit columns by analyzing balance changes.")

    bank_pdf = st.file_uploader(
        "Upload Bank Statement PDF", 
        type="pdf", 
        help="Drag and drop your statement here. Supports multi-page documents.",
        key="bank_pdf_uploader"
    )

    if bank_pdf:
        with st.spinner("Step 1: Reading PDF data..."):
            raw_content = extract_text_from_pdf(bank_pdf)
        
        if st.button("🚀 Process & Generate Excel", key="bank_process_btn"):
            with st.spinner("Step 2: Executing Intelligent Parsing..."):
                bank_data = parse_bank_statement(raw_content)
                if not bank_data:
                    st.error("No valid transaction patterns found. Please check if the PDF is a standard Bank AL Habib statement.")
                else:
                    bank_df = pd.DataFrame(bank_data)
                    bank_df = bank_df[~bank_df['Details'].str.contains("Closing Balance", na=False)]
                    
                    st.markdown(f'<div class="status-card">✅ <b>Statement Analyzed:</b> Found <b>{len(bank_df)}</b> valid rows including balance checkpoints.</div>', unsafe_allow_html=True)
                    st.write("### Data Preview")
                    st.dataframe(bank_df.head(100), use_container_width=True)
                    
                    bank_excel_file = generate_bank_excel(bank_df)
                    st.download_button(
                        label="💾 Download Structured Excel File",
                        data=bank_excel_file,
                        file_name=f"AL_Habib_Extracted_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

# --- TAB 4: TAX CALCULATOR ---
with tab4:
    st.subheader("🇵🇰 Pakistan Salary Tax Calculator (2025-2026)")
    st.markdown("Accurate tax calculation based on **FBR Slabs for Salaried Individuals (Tax Year 2025-26)**.")
    
    c1, c2 = st.columns(2)
    with c1:
        monthly_salary = st.number_input("Enter Monthly Gross Salary (PKR)", value=100000, step=5000, format="%d")
    
    annual_inc, annual_tax, monthly_tax = calculate_fbr_tax(monthly_salary)
    net_salary = monthly_salary - monthly_tax
    
    st.divider()
    res_col1, res_col2, res_col3 = st.columns(3)
    with res_col1:
        st.metric(label="Annual Tax", value=f"{annual_tax:,.0f}")
    with res_col2:
        st.metric(label="Monthly Tax Deduction", value=f"{monthly_tax:,.0f}")
    with res_col3:
        st.metric(label="Net Monthly Salary", value=f"{net_salary:,.0f}")

    st.caption("Note: Calculations are based on provided FBR Salary Slabs for Tax Year 2025-26. Rebates or adjustments are not included.")

# --- TAB 5: BLOG ---
with tab5:
    st.subheader("📰 NFP Financial Insights")
    st.write("Welcome to the NazeerFinPro blog. Here we share insights on financial management and automation.")
    
    col1, col2 = st.columns(2)
    with col1:
        try:
            st.image("https://images.unsplash.com/photo-1554224155-8d04cb21cd6c?w=800", caption="Financial Analysis")
        except:
            pass
        st.markdown("#### 5 Ways to Automate Your Payroll")
        st.write("Stop doing manual data entry. Learn how Python can save you 20 hours a week...")
        st.button("Read More", key="b1")
    with col2:
        try:
            st.image("https://images.unsplash.com/photo-1460925895917-afdab827c52f?w=800", caption="Data Visualization")
        except:
             pass
        st.markdown("#### The Power of Visual Data")
        st.write("Why your stakeholders ignore your spreadsheets and how to fix it with dashboards...")
        st.button("Read More", key="b2")

# --- TAB 6: CONTACT ---
with tab6:
    st.subheader("🤝 Work with NazeerFinPro")
    st.write("Ready to automate your financial processes? Let's connect.")
    
    c1, c2 = st.columns([1, 2])
    with c1:
        try:
            profile = PILImage.open("Nazeer Fin Pro - NFP.jpg") 
            st.image(profile, width=200)
        except:
            st.info("[Profile Photo Placeholder]")
    with c2:
        st.markdown("""
        **Nazeer Ahmed Khan** *Finance Consultant & Automation Expert*
        
        📧 **Email:** [nazeerfinpro@gmail.com](mailto:nazeerfinpro@gmail.com)  
        📱 **Phone / WhatsApp:** [+92 333 3126614](https://wa.me/923333126614)  
        🌐 **Website:** [LinkedIn Company Page](https://www.linkedin.com/company/nazeer-fin-pro)  
        """)
        st.write("---")
        st.write("**Send a Message:**")
        contact_msg = st.text_area("Message content:", placeholder="Hi Nazeer, I need help with...", label_visibility="collapsed")
        
        if st.button("Prepare Message"):
            if contact_msg:
                subject = "Inquiry from NFP Tool Suite"
                encoded_body = urllib.parse.quote(contact_msg)
                encoded_subject = urllib.parse.quote(subject)
                mailto_url = f"mailto:nazeerfinpro@gmail.com?subject={encoded_subject}&body={encoded_body}"
                whatsapp_url = f"https://wa.me/923333126614?text={encoded_body}"
                
                st.success("Message Prepared! Choose how to send it:")
                st.markdown(f"""
                <div style="display: flex; gap: 10px; margin-top: 10px;">
                    <a href="{mailto_url}" target="_blank" style="padding: 10px 20px; background-color: #003366; color: white; text-decoration: none; border-radius: 5px; font-weight: bold; border: 1px solid #002244;">📧 Send via Email</a>
                    <a href="{whatsapp_url}" target="_blank" style="padding: 10px 20px; background-color: #25D366; color: white; text-decoration: none; border-radius: 5px; font-weight: bold; border: 1px solid #128C7E;">💬 Send via WhatsApp</a>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.warning("⚠️ Please write a message first.")

# Footer
st.markdown(f"<div class='footer'>© 2025 {BRAND_NAME}. All Rights Reserved. Powered by Python & Streamlit.</div>", unsafe_allow_html=True)